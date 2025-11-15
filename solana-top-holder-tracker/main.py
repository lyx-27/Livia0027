import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import time
import os
import sys
from solana_tracker_api import get_top_holders

EXCEL_FILE_NAME = "solana_token_holders.xlsx"

def initialize_excel(token_address, excel_file_name):
    global EXCEL_FILE_NAME
    EXCEL_FILE_NAME = excel_file_name

    if os.path.exists(EXCEL_FILE_NAME):
        workbook = openpyxl.load_workbook(EXCEL_FILE_NAME)
    else:
        workbook = openpyxl.Workbook()

    if 'Token_Tracking' not in workbook.sheetnames:
        token_sheet = workbook.create_sheet('Token_Tracking', 0)
    else:
        token_sheet = workbook['Token_Tracking']
    
    if token_sheet['B1'].value is None:
        token_sheet['A1'] = 'Token Address:'
        token_sheet['B1'] = token_address

    if 'Wallet_Changes' not in workbook.sheetnames:
        changes_sheet = workbook.create_sheet('Wallet_Changes')
    else:
        changes_sheet = workbook['Wallet_Changes']

    if changes_sheet['A1'].value is None:
        changes_sheet['A1'] = 'Timestamp'
        changes_sheet['B1'] = 'Added Wallets (Percentage)'
        changes_sheet['C1'] = 'Removed Wallets (Percentage)'

    workbook.save(EXCEL_FILE_NAME)
    return workbook

def get_previous_holders(token_sheet):
    if token_sheet.max_column < 2:
        return None

    last_data_column_index = token_sheet.max_column
    previous_holders_data = []
    for row_index in range(4, token_sheet.max_row + 1):
        cell_value = token_sheet.cell(row=row_index, column=last_data_column_index).value
        if cell_value:
            lines = cell_value.split('\n')
            for line in lines:
                parts = line.strip().split(' (')
                if len(parts) == 2:
                    address = parts[0]
                    percentage_str = parts[1].replace('%', '').replace(')', '')
                    try:
                        percentage = float(percentage_str)
                        previous_holders_data.append({'address': address, 'percentage': percentage})
                    except ValueError:
                        pass
    return previous_holders_data

def write_holders_to_excel(workbook, current_holders, previous_holders=None):
    token_sheet = workbook['Token_Tracking']
    changes_sheet = workbook['Wallet_Changes']

    next_col_index = token_sheet.max_column + 1
    
    timestamp_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    token_sheet.cell(row=3, column=next_col_index, value=timestamp_str)

    holder_display_data = []
    current_addresses_with_percentage = {}
    for holder in current_holders:
        display_string = f"{holder['address']} ({holder['percentage']:.4f}%)"
        holder_display_data.append(display_string)
        current_addresses_with_percentage[holder['address']] = holder['percentage']

    for i, data_line in enumerate(holder_display_data):
        token_sheet.cell(row=i+4, column=next_col_index, value=data_line)

    if previous_holders:
        previous_addresses_with_percentage = {holder['address']: holder['percentage'] for holder in previous_holders}
        previous_addresses = set(previous_addresses_with_percentage.keys())
        current_addresses = set(current_addresses_with_percentage.keys())

        added_wallets = list(current_addresses - previous_addresses)
        removed_wallets = list(previous_addresses - current_addresses)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, holder_data in enumerate(current_holders):
            if holder_data['address'] in added_wallets:
                token_sheet.cell(row=i+4, column=next_col_index).fill = yellow_fill

        if added_wallets or removed_wallets:
            changes_row = changes_sheet.max_row + 1
            changes_sheet.cell(row=changes_row, column=1, value=timestamp_str)
            
            added_display = [f"{addr} ({current_addresses_with_percentage[addr]:.4f}%)" for addr in added_wallets]
            removed_display = [f"{addr} ({previous_addresses_with_percentage[addr]:.4f}%)" for addr in removed_wallets]

            changes_sheet.cell(row=changes_row, column=2, value="\n".join(added_display))
            changes_sheet.cell(row=changes_row, column=3, value="\n".join(removed_display))

    workbook.save(EXCEL_FILE_NAME)

def run_tracker(token_address, excel_file_name):
    workbook = initialize_excel(token_address, excel_file_name)

    print(f"Tracking token address: {token_address}, data saved to {excel_file_name}")

    previous_holders = get_previous_holders(workbook['Token_Tracking'])

    print(f"Fetching Top Holders for {token_address}...")
    current_holders_data = get_top_holders(token_address)
    
    if isinstance(current_holders_data, list):
        current_holders = current_holders_data
    elif isinstance(current_holders_data, dict) and 'data' in current_holders_data and 'items' in current_holders_data['data']:
        current_holders = current_holders_data['data']['items']
    else:
        current_holders = []

    if current_holders:
        print(f"Fetch successful, writing to Excel file.")
        write_holders_to_excel(workbook, current_holders, previous_holders)
    else:
        print("Fetch failed. Check API Key or network connection.")

    print("Tracking task finished.")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 main.py <token_address> <excel_file_name>")
        print("Example: python3 main.py EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v my_token_data.xlsx")
        sys.exit(1)

    token_address = sys.argv[1]
    excel_file_name = sys.argv[2]
    run_tracker(token_address, excel_file_name)
